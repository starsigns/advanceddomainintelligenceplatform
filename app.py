import os
import duckdb
import pandas as pd
import time
import threading
from datetime import datetime
from flask import Flask, render_template, request, redirect, url_for, send_file, flash, jsonify
import requests
from dotenv import load_dotenv
import csv
import io
import logging
from collections import defaultdict
import zipfile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import re

# Load environment variables
load_dotenv()

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', 'your-secret-key-change-this')

# API configuration
API_PROVIDER = os.getenv('API_PROVIDER', 'viewdns')
VIEWDNS_API_KEY = os.getenv('VIEWDNS_API_KEY')
SECURITYTRAILS_API_KEY = os.getenv('SECURITYTRAILS_API_KEY')

# Database configuration
DATABASE = 'domains.duckdb'

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# Global variables for tracking harvesting progress
harvest_progress = {}
harvest_threads = {}

def init_db():
    """Initialize the DuckDB database with enhanced schema for provider tracking."""
    conn = duckdb.connect(DATABASE)
    
    # Enable optimizations (using correct DuckDB syntax)
    conn.execute("SET enable_object_cache=true")
    
    # Check if domains table exists
    try:
        result = conn.execute("SELECT COUNT(*) FROM domains").fetchone()
        table_exists = True
    except:
        table_exists = False
    
    if not table_exists:
        # Create domains table with optimized types
        conn.execute('''
            CREATE TABLE domains (
                id BIGINT PRIMARY KEY,
                domain VARCHAR NOT NULL,
                mx VARCHAR,
                ns VARCHAR,
                provider VARCHAR DEFAULT 'viewdns',
                session_id VARCHAR,
                fetched_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Create sequence for auto-increment
        conn.execute("CREATE SEQUENCE domains_id_seq START 1")
        
        # Create optimized indexes
        conn.execute("CREATE INDEX idx_domain ON domains(domain)")
        conn.execute("CREATE INDEX idx_mx ON domains(mx)")
        conn.execute("CREATE INDEX idx_ns ON domains(ns)")
        conn.execute("CREATE INDEX idx_provider ON domains(provider)")
        conn.execute("CREATE INDEX idx_fetched_at ON domains(fetched_at)")
    
    # Check if harvest_sessions table exists
    try:
        result = conn.execute("SELECT COUNT(*) FROM harvest_sessions").fetchone()
        sessions_table_exists = True
    except:
        sessions_table_exists = False
    
    if not sessions_table_exists:
        # Create harvest sessions table
        conn.execute('''
            CREATE TABLE harvest_sessions (
                id VARCHAR PRIMARY KEY,
                server VARCHAR NOT NULL,
                record_type VARCHAR NOT NULL,
                provider VARCHAR NOT NULL,
                status VARCHAR DEFAULT 'running',
                total_domains BIGINT DEFAULT 0,
                pages_fetched BIGINT DEFAULT 0,
                started_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                completed_at TIMESTAMP
            )
        ''')
        
        # Create indexes for harvest sessions
        conn.execute("CREATE INDEX idx_harvest_server ON harvest_sessions(server)")
        conn.execute("CREATE INDEX idx_harvest_status ON harvest_sessions(status)")
        conn.execute("CREATE INDEX idx_harvest_started ON harvest_sessions(started_at)")
    
    conn.close()

# Global variables for tracking harvesting progress
harvest_progress = {}
harvest_threads = {}

class ViewDNSAPI:
    """ViewDNS API client with unlimited auto-scroll harvesting."""
    
    def __init__(self, api_key):
        self.api_key = api_key
        self.base_url = 'https://api.viewdns.info'
        
    def reverse_mx_lookup(self, mx_host, page=1):
        """Fetch reverse MX lookup data from ViewDNS API."""
        url = f"{self.base_url}/reversemx/"
        params = {
            'mx': mx_host,
            'apikey': self.api_key,
            'output': 'json',
            'page': page
        }
        
        try:
            response = requests.get(url, params=params, timeout=30)
            response.raise_for_status()
            
            # Check if response is JSON
            content_type = response.headers.get('content-type', '')
            if 'application/json' not in content_type:
                logger.error(f"API returned non-JSON response: {response.text[:200]}")
                return None
            
            data = response.json()
            
            # Handle API error responses
            if isinstance(data, str):
                logger.error(f"API returned error string: {data}")
                return None
                
            if 'response' not in data:
                logger.error(f"Unexpected API response structure: {data}")
                return None
                
            return data['response']
            
        except requests.exceptions.RequestException as e:
            logger.error(f"ViewDNS API request failed: {e}")
            return None
        except ValueError as e:
            logger.error(f"Failed to parse JSON response: {e}")
            return None
        except Exception as e:
            logger.error(f"ViewDNS API error: {e}")
            return None
    
    def reverse_ns_lookup(self, ns_host, page=1):
        """Fetch reverse NS lookup data from ViewDNS API."""
        url = f"{self.base_url}/reversens/"
        params = {
            'ns': ns_host,
            'apikey': self.api_key,
            'output': 'json',
            'page': page
        }
        
        try:
            response = requests.get(url, params=params, timeout=30)
            response.raise_for_status()
            
            # Check if response is JSON
            content_type = response.headers.get('content-type', '')
            if 'application/json' not in content_type:
                logger.error(f"API returned non-JSON response: {response.text[:200]}")
                return None
            
            data = response.json()
            
            # Handle API error responses
            if isinstance(data, str):
                logger.error(f"API returned error string: {data}")
                return None
                
            if 'response' not in data:
                logger.error(f"Unexpected API response structure: {data}")
                return None
                
            return data['response']
            
        except requests.exceptions.RequestException as e:
            logger.error(f"ViewDNS API request failed: {e}")
            return None
        except ValueError as e:
            logger.error(f"Failed to parse JSON response: {e}")
            return None
        except Exception as e:
            logger.error(f"ViewDNS API error: {e}")
            return None
    
    def harvest_all_domains(self, record_type, server, session_id, max_pages=None):
        """Harvest all domains using auto-scroll pagination."""
        logger.info(f"Starting harvest for {record_type.upper()} server: {server}")
        
        page = 1
        total_domains = 0
        consecutive_empty_pages = 0
        max_empty_pages = 3  # Stop after 3 consecutive empty pages
        
        # Update session status
        update_harvest_session(session_id, status='running', pages_fetched=0)
        
        while True:
            # Check max pages limit
            if max_pages and page > max_pages:
                logger.info(f"Reached max pages limit: {max_pages}")
                break
                
            logger.info(f"Fetching page {page} for {server}...")
            
            # Fetch data based on record type
            if record_type == 'mx':
                data = self.reverse_mx_lookup(server, page)
            else:
                data = self.reverse_ns_lookup(server, page)
            
            if not data or 'domains' not in data:
                consecutive_empty_pages += 1
                logger.warning(f"No data on page {page} (consecutive empty: {consecutive_empty_pages})")
                
                if consecutive_empty_pages >= max_empty_pages:
                    logger.info(f"Stopping after {consecutive_empty_pages} consecutive empty pages")
                    break
                    
                page += 1
                continue
            
            domains = data['domains']
            if not domains:
                consecutive_empty_pages += 1
                logger.warning(f"Empty domains list on page {page}")
                
                if consecutive_empty_pages >= max_empty_pages:
                    logger.info(f"Stopping after {consecutive_empty_pages} consecutive empty pages")
                    break
                    
                page += 1
                continue
            
            # Reset consecutive empty pages counter
            consecutive_empty_pages = 0
            
            # Insert domains into database
            page_count = insert_domains_batch(domains, record_type, server, session_id)
            total_domains += page_count
            
            logger.info(f"Page {page}: {page_count} domains, Total: {total_domains}")
            
            # Update session progress
            update_harvest_session(session_id, pages_fetched=page, total_domains=total_domains)
            
            # Update global progress
            harvest_progress[session_id] = {
                'page': page,
                'total_domains': total_domains,
                'server': server,
                'record_type': record_type
            }
            
            page += 1
            
            # Rate limiting - be respectful to the API
            time.sleep(1)
        
        # Mark session as complete
        update_harvest_session(session_id, status='complete', total_domains=total_domains)
        
        logger.info(f"Harvest complete for {server}: {total_domains} domains across {page-1} pages")
        return total_domains

def format_domain_url(domain):
    """
    Format domain as clickable URL by adding http:// prefix if not present.
    Ensures consistent URL formatting across the application.
    
    Args:
        domain (str): Domain name to format
        
    Returns:
        str: Formatted URL with http:// prefix
    """
    if not domain:
        return domain
    
    # Check if already has protocol
    if domain.startswith(('http://', 'https://')):
        return domain
    
    # Add http:// prefix to make clickable
    return f"http://{domain}"

def insert_domains_batch(domains, record_type, server, session_id):
    """Insert a batch of domains into the database using DuckDB."""
    conn = duckdb.connect(DATABASE)
    
    inserted_count = 0
    
    # Handle case where domains might be a string or unexpected format
    if not isinstance(domains, list):
        logger.error(f"Expected list of domains, got: {type(domains)} - {domains}")
        conn.close()
        return 0
    
    # Prepare batch data for insertion
    batch_data = []
    
    for domain in domains:
        # ViewDNS returns domains as plain strings, not objects
        if isinstance(domain, str):
            domain_name = domain.strip()
        elif isinstance(domain, dict):
            # Fallback for other APIs that might return objects
            domain_name = domain.get('domain', '').strip()
        else:
            logger.warning(f"Unexpected domain format: {domain}")
            continue
            
        if not domain_name:
            continue
            
        mx_value = server if record_type == 'mx' else None
        ns_value = server if record_type == 'ns' else None
        
        # Format domain as clickable URL
        formatted_domain = format_domain_url(domain_name)
        
        # Add to batch data
        batch_data.append((
            formatted_domain, 
            mx_value, 
            ns_value, 
            'viewdns', 
            session_id
        ))
    
    # Bulk insert using DuckDB - much faster than individual inserts
    if batch_data:
        try:
            # Create DataFrame for faster bulk insert
            df = pd.DataFrame(batch_data, columns=['domain', 'mx', 'ns', 'provider', 'session_id'])
            
            # Use DuckDB's INSERT from DataFrame (extremely fast)
            conn.execute("""
                INSERT OR IGNORE INTO domains (id, domain, mx, ns, provider, session_id)
                SELECT nextval('domains_id_seq'), domain, mx, ns, provider, session_id
                FROM df
            """)
            
            inserted_count = len(batch_data)
            logger.info(f"Bulk inserted {inserted_count} domains")
            
        except Exception as e:
            logger.error(f"Bulk insert error: {e}")
            # Fallback to individual inserts
            for data in batch_data:
                try:
                    conn.execute("""
                        INSERT OR IGNORE INTO domains (id, domain, mx, ns, provider, session_id)
                        VALUES (nextval('domains_id_seq'), ?, ?, ?, ?, ?)
                    """, data)
                    inserted_count += 1
                except Exception as inner_e:
                    logger.error(f"Individual insert error: {inner_e}")
    
    conn.close()
    return inserted_count

def create_harvest_session(server, record_type, provider='viewdns'):
    """Create a new harvest session."""
    session_id = f"{provider}_{record_type}_{server}_{int(time.time())}"
    
    conn = duckdb.connect(DATABASE)
    
    conn.execute('''
        INSERT INTO harvest_sessions (id, server, record_type, provider)
        VALUES (?, ?, ?, ?)
    ''', (session_id, server, record_type, provider))
    
    conn.close()
    
    return session_id

def update_harvest_session(session_id, **kwargs):
    """Update harvest session with new information."""
    conn = duckdb.connect(DATABASE)
    
    updates = []
    values = []
    
    for key, value in kwargs.items():
        if key in ['status', 'total_domains', 'pages_fetched']:
            updates.append(f"{key} = ?")
            values.append(value)
    
    if updates:
        if kwargs.get('status') == 'complete':
            updates.append("completed_at = CURRENT_TIMESTAMP")
            
        query = f"UPDATE harvest_sessions SET {', '.join(updates)} WHERE id = ?"
        values.append(session_id)
        
        conn.execute(query, values)
    
    conn.close()

def get_database_stats():
    """Get comprehensive database statistics with DuckDB optimizations."""
    conn = duckdb.connect(DATABASE)
    
    try:
        # Use single query with multiple aggregations for better performance
        stats_query = """
        WITH domain_stats AS (
            SELECT 
                COUNT(*) as total_records,
                COUNT(DISTINCT domain) as unique_domains,
                COUNT(CASE WHEN mx IS NOT NULL THEN 1 END) as mx_count,
                COUNT(CASE WHEN ns IS NOT NULL THEN 1 END) as ns_count
            FROM domains
        ),
        provider_stats AS (
            SELECT provider, COUNT(*) as count 
            FROM domains 
            GROUP BY provider
        ),
        top_mx AS (
            SELECT mx, COUNT(*) as count 
            FROM domains 
            WHERE mx IS NOT NULL 
            GROUP BY mx 
            ORDER BY count DESC 
            LIMIT 10
        ),
        top_ns AS (
            SELECT ns, COUNT(*) as count 
            FROM domains 
            WHERE ns IS NOT NULL 
            GROUP BY ns 
            ORDER BY count DESC 
            LIMIT 10
        ),
        recent_harvests AS (
            SELECT server, record_type, provider, total_domains, status, started_at
            FROM harvest_sessions 
            ORDER BY started_at DESC 
            LIMIT 10
        )
        SELECT * FROM domain_stats
        """
        
        # Get main stats
        result = conn.execute(stats_query).fetchone()
        total_records, unique_domains, mx_count, ns_count = result
        
        # Get provider breakdown
        by_provider = dict(conn.execute("SELECT provider, COUNT(*) FROM domains GROUP BY provider").fetchall())
        
        # Get top servers
        top_mx_servers = conn.execute("""
            SELECT mx, COUNT(*) as count FROM domains 
            WHERE mx IS NOT NULL 
            GROUP BY mx 
            ORDER BY count DESC 
            LIMIT 10
        """).fetchall()
        
        top_ns_servers = conn.execute("""
            SELECT ns, COUNT(*) as count FROM domains 
            WHERE ns IS NOT NULL 
            GROUP BY ns 
            ORDER BY count DESC 
            LIMIT 10
        """).fetchall()
        
        # Get recent harvests
        recent_harvests = conn.execute("""
            SELECT server, record_type, provider, total_domains, status, started_at
            FROM harvest_sessions 
            ORDER BY started_at DESC 
            LIMIT 10
        """).fetchall()
        
        return {
            'total_records': total_records,
            'unique_domains': unique_domains,
            'by_type': {'mx': mx_count, 'ns': ns_count},
            'by_provider': by_provider,
            'top_mx_servers': top_mx_servers,
            'top_ns_servers': top_ns_servers,
            'recent_harvests': recent_harvests
        }
        
    except Exception as e:
        logger.error(f"Error getting database stats: {e}")
        return {
            'total_records': 0,
            'unique_domains': 0,
            'by_type': {'mx': 0, 'ns': 0},
            'by_provider': {},
            'top_mx_servers': [],
            'top_ns_servers': [],
            'recent_harvests': []
        }
    finally:
        conn.close()

@app.route('/')
def home():
    """Main dashboard with harvesting interface."""
    stats = get_database_stats()
    default_provider = API_PROVIDER
    return render_template('index.html', stats=stats, default_provider=default_provider)

@app.route('/harvest', methods=['POST'])
def harvest():
    """Start a new domain harvest."""
    record_type = request.form.get('record_type')
    server = request.form.get('server', '').strip()
    api_provider = request.form.get('api_provider', 'viewdns')
    max_pages = request.form.get('max_pages')
    
    if not record_type or not server:
        flash('Please provide both record type and server name.', 'error')
        return redirect(url_for('home'))
    
    if record_type not in ['mx', 'ns']:
        flash('Invalid record type. Please select MX or NS.', 'error')
        return redirect(url_for('home'))
    
    # Convert max_pages to int if provided
    max_pages_int = None
    if max_pages:
        try:
            max_pages_int = int(max_pages)
            if max_pages_int <= 0:
                max_pages_int = None
        except ValueError:
            flash('Invalid max pages value. Using unlimited.', 'warning')
    
    # Create harvest session
    session_id = create_harvest_session(server, record_type, api_provider)
    
    if api_provider == 'viewdns':
        if not VIEWDNS_API_KEY:
            flash('ViewDNS API key not configured. Please check your .env file.', 'error')
            return redirect(url_for('home'))
        
        # Start harvest in background thread
        def harvest_worker():
            try:
                api = ViewDNSAPI(VIEWDNS_API_KEY)
                total = api.harvest_all_domains(record_type, server, session_id, max_pages_int)
                logger.info(f'Harvest completed! Fetched {total:,} domains for {server}')
            except Exception as e:
                logger.error(f"Harvest error: {e}")
                update_harvest_session(session_id, status='error')
        
        thread = threading.Thread(target=harvest_worker)
        thread.daemon = True
        thread.start()
        harvest_threads[session_id] = thread
        
        flash(f'Started harvesting {record_type.upper()} domains for {server} using ViewDNS API!', 'info')
        
    else:
        flash('SecurityTrails integration coming soon. Please use ViewDNS for now.', 'warning')
    
    return redirect(url_for('home'))

def sanitize_filename(filename):
    """Sanitize filename to remove invalid characters."""
    # Remove or replace invalid characters
    filename = re.sub(r'[<>:"/\\|?*]', '_', filename)
    # Remove multiple consecutive underscores
    filename = re.sub(r'_+', '_', filename)
    # Remove leading/trailing underscores and dots
    filename = filename.strip('_.')
    # Limit length
    if len(filename) > 200:
        filename = filename[:200]
    return filename

def generate_filename(server, record_type, file_format, timestamp=None):
    """Generate a filename based on server, record type, and timestamp."""
    if timestamp is None:
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    # Sanitize server name for filename
    safe_server = sanitize_filename(server) if server else 'all_servers'
    safe_record_type = record_type if record_type in ['mx', 'ns'] else 'all_types'
    
    filename = f"{safe_server}_{safe_record_type}_{timestamp}.{file_format}"
    return filename

def create_excel_export(results, server=None, record_type=None):
    """Create an Excel file with the results - handles large datasets by chunking."""
    total_results = len(results)
    max_excel_rows = 1048575  # Excel limit minus header row
    
    # If dataset is small enough for single Excel file
    if total_results <= max_excel_rows:
        wb = Workbook()
        ws = wb.active
        
        # Set sheet name
        sheet_name = f"{record_type.upper() if record_type else 'ALL'}_Records"
        ws.title = sheet_name
        
        # Define headers
        headers = ['Domain', 'MX Record', 'NS Record', 'Provider', 'Fetched At']
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add data
        for row, record in enumerate(results, 2):
            ws.cell(row=row, column=1, value=record[0])  # Domain
            ws.cell(row=row, column=2, value=record[1] or '')  # MX
            ws.cell(row=row, column=3, value=record[2] or '')  # NS
            ws.cell(row=row, column=4, value=record[3] or 'Unknown')  # Provider
            ws.cell(row=row, column=5, value=record[4])  # Fetched At
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Max width of 50
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Add summary sheet if large dataset
        if len(results) > 1000:
            summary_ws = wb.create_sheet("Summary")
            summary_data = [
                ["Export Summary", ""],
                ["Total Records", len(results)],
                ["Server", server or "All Servers"],
                ["Record Type", record_type.upper() if record_type else "All Types"],
                ["Export Date", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
                ["", ""],
                ["Top Domains", "Count"]
            ]
            
            for row, data in enumerate(summary_data, 1):
                summary_ws.cell(row=row, column=1, value=data[0])
                summary_ws.cell(row=row, column=2, value=data[1])
        
        # Save to BytesIO
        excel_file = io.BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)
        
        return excel_file
    
    # For large datasets, create multiple Excel files in a ZIP
    else:
        return create_chunked_excel_export(results, server, record_type)

def create_chunked_excel_export(results, server=None, record_type=None):
    """Create a single Excel file with multiple worksheets for large datasets."""
    max_excel_rows = 1048575  # Excel limit minus header row
    headers = ['Domain', 'MX Record', 'NS Record', 'Provider', 'Fetched At']
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)
    
    # Style definitions
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Create summary sheet first
    summary_ws = wb.create_sheet("Summary", 0)
    total_sheets = (len(results) // max_excel_rows) + 1
    
    summary_data = [
        ["Export Summary", ""],
        ["Total Records", len(results)],
        ["Total Worksheets", total_sheets],
        ["Server", server or "All Servers"],
        ["Record Type", record_type.upper() if record_type else "All Types"],
        ["Export Date", datetime.now().strftime('%Y-%m-%d %H:%M:%S')],
        ["", ""],
        ["Worksheet", "Records", "Range"],
    ]
    
    # Add worksheet breakdown
    for i in range(0, len(results), max_excel_rows):
        sheet_num = (i // max_excel_rows) + 1
        chunk_size = min(max_excel_rows, len(results) - i)
        start_record = i + 1
        end_record = i + chunk_size
        summary_data.append([f"Data_Sheet_{sheet_num}", chunk_size, f"{start_record}-{end_record}"])
    
    # Populate summary sheet
    for row, data in enumerate(summary_data, 1):
        for col, value in enumerate(data, 1):
            cell = summary_ws.cell(row=row, column=col, value=value)
            if row == 1:  # Header row
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            elif row == 8:  # Column headers
                cell.font = Font(bold=True)
    
    # Auto-adjust summary column widths
    for column in summary_ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        summary_ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create data sheets
    for i in range(0, len(results), max_excel_rows):
        chunk = results[i:i + max_excel_rows]
        sheet_num = (i // max_excel_rows) + 1
        
        # Create worksheet
        sheet_name = f"Data_Sheet_{sheet_num}"
        ws = wb.create_sheet(sheet_name)
        
        # Add headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Add data
        for row, record in enumerate(chunk, 2):
            ws.cell(row=row, column=1, value=record[0])  # Domain
            ws.cell(row=row, column=2, value=record[1] or '')  # MX
            ws.cell(row=row, column=3, value=record[2] or '')  # NS
            ws.cell(row=row, column=4, value=record[3] or 'Unknown')  # Provider
            ws.cell(row=row, column=5, value=record[4])  # Fetched At
        
        # Auto-adjust column widths (sample first 100 rows for performance)
        sample_rows = min(100, len(chunk))
        for col_idx, column in enumerate(ws.columns, 1):
            max_length = len(headers[col_idx - 1])  # Start with header length
            for row_idx in range(2, sample_rows + 2):  # Sample rows
                cell = ws.cell(row=row_idx, column=col_idx)
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Max width of 50
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    # Save to BytesIO
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return excel_file

def create_chunked_csv_export(results, server=None, record_type=None, chunk_size=50000):
    """Create multiple CSV files if dataset is too large."""
    if len(results) <= chunk_size:
        return create_single_csv_export(results)
    
    # Create ZIP file with multiple CSV chunks
    zip_buffer = io.BytesIO()
    
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
        headers = ['Domain', 'MX Record', 'NS Record', 'Provider', 'Fetched At']
        
        for i in range(0, len(results), chunk_size):
            chunk = results[i:i + chunk_size]
            chunk_num = (i // chunk_size) + 1
            
            # Create CSV for this chunk
            csv_buffer = io.StringIO()
            writer = csv.writer(csv_buffer)
            writer.writerow(headers)
            writer.writerows(chunk)
            
            # Add to ZIP
            chunk_filename = f"chunk_{chunk_num:03d}.csv"
            zip_file.writestr(chunk_filename, csv_buffer.getvalue())
    
    zip_buffer.seek(0)
    return zip_buffer

def create_single_csv_export(results):
    """Create a single CSV file."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['Domain', 'MX Record', 'NS Record', 'Provider', 'Fetched At'])
    writer.writerows(results)
    
    csv_data = output.getvalue()
    output.close()
    
    csv_file = io.BytesIO(csv_data.encode('utf-8'))
    csv_file.seek(0)
    return csv_file

@app.route('/export')
def export():
    """Export domains with multiple format options."""
    record_type = request.args.get('type')
    server = request.args.get('server')
    format_type = request.args.get('format', 'excel')  # Default to Excel
    
    conn = duckdb.connect(DATABASE)
    
    # Build query based on filters
    query = "SELECT domain, mx, ns, provider, fetched_at FROM domains WHERE 1=1"
    params = []
    
    if record_type == 'mx':
        query += " AND mx IS NOT NULL"
        if server:
            query += " AND mx = ?"
            params.append(server)
    elif record_type == 'ns':
        query += " AND ns IS NOT NULL"
        if server:
            query += " AND ns = ?"
            params.append(server)
    
    query += " ORDER BY fetched_at DESC"
    
    # Use pandas for faster data processing
    df = conn.execute(query, params).df()
    results = df.to_records(index=False).tolist()
    conn.close()
    
    if not results:
        flash('No data to export!', 'warning')
        return redirect(url_for('home'))
    
    # Generate timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    # Handle different export formats
    if format_type == 'excel':
        file_buffer = create_excel_export(results, server, record_type)
        filename = generate_filename(server, record_type, 'xlsx', timestamp)
        mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        
    elif format_type == 'csv_chunked' or len(results) > 100000:
        file_buffer = create_chunked_csv_export(results, server, record_type)
        filename = generate_filename(server, record_type, 'zip', timestamp)
        mimetype = 'application/zip'
        
    else:  # Single CSV
        file_buffer = create_single_csv_export(results)
        filename = generate_filename(server, record_type, 'csv', timestamp)
        mimetype = 'text/csv'
    
    return send_file(
        file_buffer,
        as_attachment=True,
        download_name=filename,
        mimetype=mimetype
    )

@app.route('/api/stats')
def api_stats():
    """API endpoint for live stats."""
    stats = get_database_stats()
    return jsonify(stats)

@app.route('/test_viewdns')
def test_viewdns():
    """Test ViewDNS API response format."""
    if not VIEWDNS_API_KEY:
        return jsonify({'error': 'ViewDNS API key not configured'})
    
    api = ViewDNSAPI(VIEWDNS_API_KEY)
    
    # Test with a known working server
    test_server = 'mx01.ionos.de'
    logger.info(f"Testing ViewDNS API with {test_server}")
    
    try:
        # Make raw API request to see response format
        url = f"{api.base_url}/reversemx/"
        params = {
            'mx': test_server,
            'apikey': api.api_key,
            'output': 'json',
            'page': 1
        }
        
        response = requests.get(url, params=params, timeout=30)
        logger.info(f"Raw response status: {response.status_code}")
        logger.info(f"Raw response headers: {dict(response.headers)}")
        logger.info(f"Raw response text (first 500 chars): {response.text[:500]}")
        
        # Try to parse as JSON
        try:
            data = response.json()
            logger.info(f"Parsed JSON structure: {type(data)} - Keys: {list(data.keys()) if isinstance(data, dict) else 'Not a dict'}")
            return jsonify({
                'success': True,
                'status_code': response.status_code,
                'content_type': response.headers.get('content-type'),
                'data_type': str(type(data)),
                'data_keys': list(data.keys()) if isinstance(data, dict) else None,
                'raw_response': response.text[:1000]
            })
        except ValueError as e:
            return jsonify({
                'success': False,
                'error': f'JSON parse error: {e}',
                'raw_response': response.text[:1000]
            })
            
    except Exception as e:
        logger.error(f"ViewDNS test error: {e}")
        return jsonify({'error': str(e)})

@app.route('/progress/<session_id>')
def progress(session_id):
    """Get harvest progress for a session."""
    progress_data = harvest_progress.get(session_id, {})
    return jsonify(progress_data)

@app.route('/clear-database', methods=['POST'])
def clear_database():
    """Safely clear all data from the database while preserving structure."""
    try:
        conn = duckdb.connect(DATABASE)
        
        # Get counts before clearing
        domains_count = conn.execute('SELECT COUNT(*) FROM domains').fetchone()[0]
        sessions_count = conn.execute('SELECT COUNT(*) FROM harvest_sessions').fetchone()[0]
        
        # Clear all data from tables (but keep the structure)
        conn.execute('DELETE FROM domains')
        conn.execute('DELETE FROM harvest_sessions')
        
        # Reset sequence for DuckDB (DuckDB uses different syntax)
        try:
            conn.execute('DROP SEQUENCE IF EXISTS domains_id_seq')
            conn.execute('CREATE SEQUENCE domains_id_seq START 1')
        except Exception as seq_error:
            # If sequence operations fail, just continue - the main clearing still worked
            logger.warning(f"Could not reset sequence: {seq_error}")
        
        conn.close()
        
        # Clear any active harvest progress tracking
        global harvest_progress, harvest_threads
        harvest_progress.clear()
        harvest_threads.clear()
        
        logger.info(f"Database cleared successfully. Removed {domains_count} domains and {sessions_count} sessions")
        
        return jsonify({
            'success': True,
            'message': f'Database cleared successfully! Removed {domains_count} domains and {sessions_count} harvest sessions.',
            'domains_removed': domains_count,
            'sessions_removed': sessions_count
        })
        
    except Exception as e:
        logger.error(f"Error clearing database: {e}")
        return jsonify({
            'success': False,
            'error': f'Failed to clear database: {str(e)}'
        }), 500

if __name__ == '__main__':
    init_db()
    app.run(debug=True, host='0.0.0.0', port=5000)
